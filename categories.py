from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json

def get_name_from_item_dict(item):
	if item['_type'] == 'Node':
		return item['_name']
	else:
		return item['_props']['Name']

def categorize():
	# Load sheet
	wb = load_workbook(filename = 'Tarkov item tracer prices.xlsx')
	sheet = wb.active

	# Load item data
	with open('bsg-data.json', encoding="utf8") as f:
	  jsonData = json.load(f)

	# Build name lookup table since spreadsheet is not keyed on uid
	nameToJson = {}
	for item in jsonData.values():
		nameToJson[get_name_from_item_dict(item)] = item

	# Constants
	FIRSTCATEGORYCOLUMN = 'J'
	REDFILL = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
	i = 2 # Skip title row
	while True:
		name = sheet['A' + str(i)].value
		if not name:
			break
		if nameToJson.get(name) == None:
			sheet['A'+ str(i)].fill = REDFILL # highlight mismatched names
		else:
			# build hierarchy for each item
			hierarchy = []
			current = nameToJson.get(name)
			while current != None:
				current = jsonData.get(current["_parent"])
				if current != None:
					hierarchy.append(get_name_from_item_dict(current))

			# insert hierarchy into category columns
			currentColumn = FIRSTCATEGORYCOLUMN
			while hierarchy:
				currName = hierarchy.pop()
				sheet[currentColumn + str(i)] = currName
				currentColumn = chr(ord(currentColumn) + 1)
		i += 1
	wb.save('Tarkov item tracer prices - Categorized.xlsx')

categorize()